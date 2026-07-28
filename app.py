# -*- coding: utf-8 -*-
"""
app.py
Streamlit aplikace: PDF výpis(y) z Nahlížení do katastru / z katastru
nemovitostí -> Excel pro hromadnou korespondenci.

Podporuje dva typy vstupních PDF (rozpozná se automaticky):
1) Jednodušší "Informace o stavbě" (nebo o pozemku) s sekcí
   "Vlastníci, jiní oprávnění" - parser.py
2) Kompletní "VÝPIS Z KATASTRU NEMOVITOSTÍ" s částí A (vlastníci) a
   částí B (bytové jednotky) - lv_parser.py

Lze nahrát i VÍCE PDF najednou (i namíchaně oba typy) - výsledek se
spojí do JEDNOHO Excel souboru (pokud jsou nahrané oba typy, budou ve
výsledném sešitu dva listy).

Spuštění:
    streamlit run app.py
"""

import io

import pandas as pd
import pdfplumber
import streamlit as st
from openpyxl.styles import PatternFill

from parser import process_pdf_to_rows, COLUMNS
from lv_parser import process_lv_pdf_to_rows, LV_COLUMNS, is_lv_document
from couple_merge import mark_married_couples, build_combined_sjm_rows

GRAY_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

st.set_page_config(
    page_title="Katastr → Excel pro hromadnou korespondenci",
    page_icon="📄",
    layout="wide",
)


def check_password() -> bool:
    """
    Jednoduchá ochrana heslem. Heslo se nastavuje v Streamlit "Secrets"
    (klíč APP_PASSWORD) - viz DEPLOY.md. Heslo se nikde neukládá do kódu
    ani do repozitáře.
    """

    def password_entered():
        correct = st.secrets.get("APP_PASSWORD", None)
        if correct is None:
            st.session_state["password_correct"] = True
            return
        if st.session_state.get("password_input", "") == correct:
            st.session_state["password_correct"] = True
            st.session_state["password_input"] = ""
        else:
            st.session_state["password_correct"] = False

    if st.session_state.get("password_correct", False):
        return True

    st.title("🔒 Přístup chráněn heslem")
    st.text_input(
        "Zadejte heslo", type="password", key="password_input",
        on_change=password_entered,
    )
    if "password_correct" in st.session_state and not st.session_state["password_correct"]:
        st.error("Nesprávné heslo, zkuste to znovu.")
    return False


if not check_password():
    st.stop()

st.title("📄 Katastr nemovitostí → Excel pro hromadnou korespondenci")
st.write(
    "Nahrajte jedno nebo víc PDF z katastru nemovitostí. Aplikace u "
    "každého automaticky pozná, o jaký typ výpisu jde:\n"
    "- **Informace o stavbě / o pozemku** (sekce „Vlastníci, jiní oprávnění“), nebo\n"
    "- **Kompletní výpis z katastru (list vlastnictví)** s částí A "
    "(vlastníci) a částí B (bytové jednotky) - výstup pak obsahuje "
    "u každého vlastníka i číslo jeho bytové jednotky.\n\n"
    "Pokud nahrajete víc souborů, spojí se do **jednoho** Excelu "
    "(pokud jsou mezi nimi oba typy dokumentů, budou ve výsledném "
    "sešitu dva listy)."
)

with st.expander("ℹ️ Co aplikace dělá a jak číst výsledný Excel (rozklikněte)"):
    st.markdown(
        """
**Co aplikace dělá:**
1. Přečte nahrané PDF a najde v něm seznam vlastníků (a u kompletního
   výpisu i jejich přiřazení k jednotlivým bytovým jednotkám).
2. Rozdělí každé jméno na oslovení, titul, jméno a příjmení a adresu
   na ulici, obec a PSČ.
3. U mužů a žen odhadne oslovení a skloní příjmení do 5. pádu
   („Vážený pane Nováku“, „Vážená paní Nováková“).
4. Rozpozná manžele na stejné adrese a u skutečných párů se zápisem
   „SJ“ v PDF vytvoří navíc společný řádek pro hromadnou korespondenci
   („Vážení manželé Novákovi“).
5. Vše vyexportuje do jednoho stažitelného Excelu.

**Jak číst výsledný Excel - sloupce:**
- **Bytová jednotka** *(jen u kompletního výpisu)* - číslo jednotky,
  ke které vlastník patří.
- **Oslovení** - hotový pozdrav včetně skloněného příjmení, přímo
  použitelný jako oslovení v dopise.
- **Jméno / Příjmení (/ Název) / Titul** - rozdělené jméno vlastníka;
  u firem a organizací je jméno prázdné a název je v poli
  „Příjmení / Název“.
- **Ulice / Obec / PSČ** *(+ Číslo domu u jednoduššího typu PDF)* -
  rozdělená adresa, připravená k použití v adresním štítku.
- **Kontrola** - **ANO** znamená, že si aplikace není jistá (např.
  neobvyklé jméno, adresa bez uvedené městské části, zahraniční
  adresa, firma) - takové řádky si před odesláním zkontrolujte podle
  sloupce **Poznámka**, kde je vysvětlen důvod.
- **Původní adresa** - adresa přesně tak, jak byla v PDF, pro ověření.
- **Pár (manželé)** / **Odeslat dopis** - pokud aplikace na stejné
  adrese najde dvojici opačného pohlaví se stejným kořenem příjmení,
  označí je a u druhého z páru nastaví „Odeslat dopis“ na NE (stačí
  poslat jeden dopis).
- U skutečných manželských párů se zápisem „SJ“ v PDF navíc přibude
  **třetí řádek** se společným oslovením („Vážení manželé Novákovi“) -
  ve staženém Excelu jsou všechny tři související řádky **podbarvené
  šedě**, abyste je snadno našli a zkontrolovali.

Podrobnější popis všech pravidel a omezení najdete v souboru
`README.md` v repozitáři aplikace.
        """
    )

for key, default in [
    ("results", None),  # list of dicts: {filename, mode, rows, debug}
]:
    if key not in st.session_state:
        st.session_state[key] = default

uploaded_files = st.file_uploader(
    "Nahrát PDF soubor(y)", type=["pdf"], accept_multiple_files=True
)

col1, _ = st.columns([1, 4])
with col1:
    process_clicked = st.button(
        "🔄 Zpracovat", type="primary", disabled=not uploaded_files
    )

if process_clicked and uploaded_files:
    results = []
    progress = st.progress(0.0, text="Zpracovávám soubory…")
    for i, uploaded_file in enumerate(uploaded_files):
        try:
            uploaded_file.seek(0)
            with pdfplumber.open(uploaded_file) as pdf:
                probe_text = "\n".join((p.extract_text() or "") for p in pdf.pages)
            lv_mode = is_lv_document(probe_text)

            uploaded_file.seek(0)
            if lv_mode:
                rows, debug = process_lv_pdf_to_rows(uploaded_file)
                results.append({
                    "filename": uploaded_file.name,
                    "mode": "lv",
                    "rows": rows,
                    "sjm_flags": debug.get("sjm_flags", [False] * len(rows)),
                    "debug": debug,
                    "error": None,
                })
            else:
                rows, full_text, section_text, sjm_flags = process_pdf_to_rows(uploaded_file)
                results.append({
                    "filename": uploaded_file.name,
                    "mode": "simple",
                    "rows": rows,
                    "sjm_flags": sjm_flags,
                    "debug": {"full_text": full_text, "section_text": section_text},
                    "error": None,
                })
        except Exception as exc:  # noqa: BLE001
            results.append({
                "filename": uploaded_file.name,
                "mode": None,
                "rows": [],
                "debug": None,
                "error": str(exc),
            })
        progress.progress((i + 1) / len(uploaded_files), text=f"Zpracováno {i + 1}/{len(uploaded_files)}")
    progress.empty()

    st.session_state["results"] = results

    n_files = len(results)
    n_errors = sum(1 for r in results if r["error"])
    n_simple = sum(1 for r in results if r["mode"] == "simple")
    n_lv = sum(1 for r in results if r["mode"] == "lv")
    total_rows = sum(len(r["rows"]) for r in results)

    if n_errors:
        for r in results:
            if r["error"]:
                st.error(f"Chyba při zpracování „{r['filename']}“: {r['error']}")

    if total_rows == 0 and n_errors < n_files:
        st.warning(
            "Ze zpracovaných souborů se nepodařilo rozpoznat žádného "
            "vlastníka. Zkontrolujte „Debug“ níže u jednotlivých souborů."
        )
    elif total_rows > 0:
        summary_bits = []
        if n_simple:
            summary_bits.append(f"{n_simple}× informace o stavbě/pozemku")
        if n_lv:
            summary_bits.append(f"{n_lv}× kompletní výpis (LV)")
        st.success(
            f"Hotovo – zpracováno {n_files - n_errors} souborů "
            f"({', '.join(summary_bits)}), celkem {total_rows} řádků."
        )

results = st.session_state.get("results")

if results and any(r["rows"] for r in results):
    n_files = len(results)
    multi_file = n_files > 1

    simple_rows = []
    lv_rows = []
    for r in results:
        flags = r.get("sjm_flags") or [False] * len(r["rows"])
        for row, is_sjm in zip(r["rows"], flags):
            row_copy = dict(row)
            if multi_file:
                row_copy["Zdrojový soubor"] = r["filename"]
            row_copy["_sjm"] = bool(is_sjm)
            if r["mode"] == "simple":
                simple_rows.append(row_copy)
            elif r["mode"] == "lv":
                lv_rows.append(row_copy)

    buffer = io.BytesIO()
    sheets_written = []

    tabs_needed = []
    if simple_rows:
        tabs_needed.append("simple")
    if lv_rows:
        tabs_needed.append("lv")

    tab_objs = st.tabs(
        ["📋 Vlastníci" if t == "simple" else "🏠 Bytové jednotky" for t in tabs_needed]
    ) if len(tabs_needed) > 1 else [st.container()]

    edited_dfs = {}

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for tab_idx, mode in enumerate(tabs_needed):
            with tab_objs[tab_idx]:
                if mode == "simple":
                    columns = COLUMNS + (["Zdrojový soubor"] if multi_file else []) + ["_sjm"]
                    df = pd.DataFrame(simple_rows, columns=columns)
                    df = mark_married_couples(
                        df,
                        prijmeni_col="Příjmení / Název",
                        osloveni_col="Oslovení",
                        address_cols=["Ulice", "Číslo domu", "Obec", "PSČ"],
                        file_col="Zdrojový soubor" if multi_file else None,
                    )
                    df, highlight_idx = build_combined_sjm_rows(
                        df,
                        jmeno_col="Jméno",
                        prijmeni_col="Příjmení / Název",
                        osloveni_col="Oslovení",
                        address_cols=["Ulice", "Číslo domu", "Obec", "PSČ"],
                        sjm_flag_col="_sjm",
                        file_col="Zdrojový soubor" if multi_file else None,
                    )
                    df = df.drop(columns=["_sjm"])
                    sheet_name = "Vlastnici"
                else:
                    columns = LV_COLUMNS + (["Zdrojový soubor"] if multi_file else []) + ["_sjm"]
                    df = pd.DataFrame(lv_rows, columns=columns)
                    df = mark_married_couples(
                        df,
                        prijmeni_col="Příjmení",
                        osloveni_col="Oslovení",
                        address_cols=["Ulice", "Obec", "PSČ"],
                        unit_col="Bytová jednotka",
                        file_col="Zdrojový soubor" if multi_file else None,
                    )
                    df, highlight_idx = build_combined_sjm_rows(
                        df,
                        jmeno_col="Jméno",
                        prijmeni_col="Příjmení",
                        osloveni_col="Oslovení",
                        address_cols=["Ulice", "Obec", "PSČ"],
                        sjm_flag_col="_sjm",
                        unit_col="Bytová jednotka",
                        file_col="Zdrojový soubor" if multi_file else None,
                    )
                    df = df.drop(columns=["_sjm"])
                    sheet_name = "Jednotky"

                n_kontrola = (df["Kontrola"] == "ANO").sum()
                n_parcount = (df["Odeslat dopis"] != "ANO").sum()
                n_sjm_combined = len(highlight_idx) // 3

                st.subheader("Náhled tabulky")
                st.caption(
                    "⚠️ Tabulka níže má v pravém horním rohu (po najetí "
                    "myší) svoji vlastní malou ikonku ke stažení – ta vždy "
                    "stáhne soubor jako **.csv** (špatně se otevírá v "
                    "Excelu). Pro správný Excel soubor použijte modré "
                    "tlačítko **„⬇️ Stáhnout Excel (.xlsx)“** níže."
                )
                if n_sjm_combined:
                    st.info(
                        f"👫 U {n_sjm_combined} manželských párů (rozpoznaných podle "
                        "zápisu „SJ“ v PDF, shodného příjmení a shodné adresy) byl "
                        "vytvořen řádek navíc se společným oslovením pro hromadnou "
                        "korespondenci. Tyto řádky (2 původní + 1 nový) jsou ve "
                        "staženém Excelu podbarvené šedě – zkontrolujte prosím, že "
                        "vypadají správně, a případně jeden z původních dvou řádků "
                        "před odesláním smažte."
                    )
                if n_parcount:
                    st.info(
                        f"💌 Nalezeno {n_parcount} řádků, kde manžel/"
                        "manželka bydlí na stejné adrese - sloupec "
                        "„Odeslat dopis“ je u nich nastaven na NE, ať pro "
                        "hromadnou korespondenci pošlete jen jeden dopis "
                        "na pár. Zkontrolujte prosím, že se páry "
                        "rozpoznaly správně."
                    )
                if n_kontrola:
                    st.info(
                        f"⚠️ {n_kontrola} z {len(df)} řádků má Kontrola = "
                        "ANO – doporučujeme je před odesláním ručně "
                        "zkontrolovat (viz sloupec Poznámka). Tabulku "
                        "níže můžete přímo opravit."
                    )

                edited_df = st.data_editor(
                    df,
                    num_rows="dynamic",
                    use_container_width=True,
                    column_config={
                        "Kontrola": st.column_config.SelectboxColumn(
                            "Kontrola", options=["ANO", "NE"]
                        ),
                        "Pár (manželé)": st.column_config.SelectboxColumn(
                            "Pár (manželé)", options=["ANO", "NE"]
                        ),
                        "Odeslat dopis": st.column_config.SelectboxColumn(
                            "Odeslat dopis", options=["ANO", "NE (viz manžel/manželka výše)"]
                        ),
                    },
                    key=f"editor_{mode}",
                )
                edited_dfs[mode] = (edited_df, sheet_name, highlight_idx)

                with st.expander(f"🔍 Debug – {('všechny soubory' if not multi_file else 'soubory')}"):
                    for r in results:
                        if r["mode"] != mode or not r.get("debug"):
                            continue
                        st.markdown(f"**{r['filename']}**")
                        debug = r["debug"]
                        if mode == "lv":
                            st.write(
                                f"Záznamů v části A: {debug.get('part_a_entries_count', '?')} | "
                                f"Jednotek celkem: {debug.get('units_total', '?')} | "
                                f"z toho typu „byt“: {debug.get('units_byt', '?')}"
                            )
                            st.text_area(
                                "Text části B (jednotky)",
                                debug.get("part_b_text", ""),
                                height=200,
                                key=f"debug_partb_{r['filename']}",
                            )
                        else:
                            st.text_area(
                                "Nalezená sekce „Vlastníci, jiní oprávnění“",
                                debug.get("section_text", ""),
                                height=200,
                                key=f"debug_section_{r['filename']}",
                            )
                        st.text_area(
                            "Celý text načtený z PDF",
                            debug.get("full_text", ""),
                            height=200,
                            key=f"debug_full_{r['filename']}",
                        )

        for mode in tabs_needed:
            edited_df, sheet_name, highlight_idx = edited_dfs[mode]
            edited_df.to_excel(writer, index=False, sheet_name=sheet_name)
            worksheet = writer.sheets[sheet_name]
            for column_cells in worksheet.columns:
                max_len = max(
                    (len(str(cell.value)) for cell in column_cells if cell.value),
                    default=0,
                )
                worksheet.column_dimensions[column_cells[0].column_letter].width = min(
                    max(max_len + 2, 12), 45
                )
            n_cols = len(edited_df.columns)
            for row_pos in highlight_idx:
                if row_pos >= len(edited_df):
                    continue
                excel_row = row_pos + 2  # +1 hlavička, +1 1-indexování
                for col_idx in range(1, n_cols + 1):
                    worksheet.cell(row=excel_row, column=col_idx).fill = GRAY_FILL
            sheets_written.append(sheet_name)

    buffer.seek(0)

    if len(sheets_written) > 1:
        file_name = "katastr_hromadna_korespondence.xlsx"
    elif sheets_written == ["Jednotky"]:
        file_name = "byty_hromadna_korespondence.xlsx"
    else:
        file_name = "vlastnici_hromadna_korespondence.xlsx"

    st.download_button(
        label="⬇️ Stáhnout Excel (.xlsx)",
        data=buffer,
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

elif not uploaded_files:
    st.info("Nahrajte jedno nebo víc PDF a klikněte na „Zpracovat“.")
else:
    st.info("Klikněte na „Zpracovat“ pro zpracování nahraných PDF.")
